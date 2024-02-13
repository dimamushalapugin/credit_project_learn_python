import os
from openpyxl import load_workbook
from openpyxl import Workbook

from webapp.config import ABS_PATH_APP, ABS_PATH_TEMP


def merge_files(file_name, download_name):
    file_path1 = os.path.join(ABS_PATH_TEMP, file_name)
    file_path2 = os.path.join(ABS_PATH_APP, 'ШАБЛОН ЗАЯВКИ ЛКМБ.xlsx')
    merged_file_path = os.path.join(ABS_PATH_TEMP, download_name)

    try:
        # Загрузка первого файла для чтения
        workbook1 = load_workbook(filename=file_path1)
        # Загрузка второго файла (если он существует) или создание нового
        try:
            workbook2 = load_workbook(filename=file_path2)
        except FileNotFoundError:
            workbook2 = Workbook()

        # Получение листа из первого файла
        sheet1 = workbook1.active
        # Копирование листа во второй файл
        sheet2 = workbook2.copy_worksheet(sheet1)
        # Переименование скопированного листа (если требуется)
        sheet2.title = 'Sheet1'

        # Сохранение изменений во втором файле
        workbook2.save(merged_file_path)

    except Exception as e:
        print(f"Ошибка при объединении файлов: {e}")
