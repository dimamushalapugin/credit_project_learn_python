import datetime
import openpyxl

from openpyxl.styles import Font, PatternFill
from datetime import datetime as dt
from dateutil.relativedelta import relativedelta

from webapp.risk.logger import logging


def main_conditions(file_name):
    wb = openpyxl.load_workbook(file_name)
    first_list(wb, file_name)
    second_list(wb, file_name)
    third_list(wb, file_name)
    # fourth_list(wb, file_name)


def check_date_of_registration(sheet):
    font_red = Font(name='Century Gothic', color='ff4444')
    font_green = Font(name='Century Gothic', color='009900')
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Дата регистрации':
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                try:
                    if not isinstance(adjacent_cell.value, datetime.datetime):
                        reg_comp_seller = dt.strptime(adjacent_cell.value, '%d.%m.%Y')
                    else:
                        reg_comp_seller = adjacent_cell.value
                    cur_date_seller = dt.strptime(dt.now().strftime('%d.%m.%Y'), '%d.%m.%Y')
                    if relativedelta(cur_date_seller, reg_comp_seller).years >= 3:
                        adjacent_cell.font = font_green
                    else:
                        adjacent_cell.font = font_red
                except Exception as _ex:
                    logging.info(_ex, exc_info=True)
                    pass


def check_size_of_authorized_capital(sheet):
    font_red = Font(name='Century Gothic', color='ff4444')
    font_green = Font(name='Century Gothic', color='009900')
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'Уставный капитал':
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                try:
                    capital = float(adjacent_cell.value.replace(' руб.', '').replace(' ', ''))
                    if capital >= 100_000:
                        adjacent_cell.font = font_green
                    else:
                        adjacent_cell.font = font_red
                except Exception as _ex:
                    logging.info(_ex, exc_info=True)
                    pass


def delta_rating(sheet):
    color_green = PatternFill(start_color='a7cfa7', end_color='a7cfa7', fill_type='solid')
    color_orange = PatternFill(start_color='fabb7c', end_color='fabb7c', fill_type='solid')
    color_red = PatternFill(start_color='fa867c', end_color='fa867c', fill_type='solid')
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == '2. Анализ деятельности':
                next_cell = sheet.cell(row=cell.row + 1, column=cell.column)
                try:
                    if int(next_cell.value) >= 70:
                        next_cell.fill = color_green
                    elif int(next_cell.value) in [50, 60]:
                        next_cell.fill = color_orange
                    else:
                        next_cell.fill = color_red
                except Exception as _ex:
                    print(_ex)
                    logging.info(_ex, exc_info=True)
                    pass


def activity_analysis(sheet):
    font_red = Font(name='Century Gothic', color='ff4444')
    font_green = Font(name='Century Gothic', color='009900')
    font_orange = Font(name='Century Gothic', color='ff8358')
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value in ['Дисквалифицированные лица:', 'Сообщения о банкротстве:', 'Арбитражные дела:']:
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                if 'отсутству' in adjacent_cell.value:
                    adjacent_cell.font = font_green
                else:
                    adjacent_cell.font = font_red
            if cell.value == 'Массовость директоров/учредителей:':
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                if 'не является' in adjacent_cell.value:
                    adjacent_cell.font = font_green
                else:
                    adjacent_cell.font = font_red
            if cell.value == 'Преступления, связанные с оборотом наркотиков по данным УФСКН России:':
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                if 'нет лиц' in adjacent_cell.value:
                    adjacent_cell.font = font_green
                else:
                    adjacent_cell.font = font_red
            if cell.value == 'Госконтракты:':
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                if 'не принимала' in adjacent_cell.value:
                    adjacent_cell.font = font_orange
                else:
                    adjacent_cell.font = font_green
            if cell.value == 'Налоговая задолженность и отчетность:':
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                if 'не имеет' in adjacent_cell.value:
                    adjacent_cell.font = font_green
                else:
                    adjacent_cell.font = font_red
            if cell.value == 'Заблокированные расчетные счета:':
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                if 'нет' in adjacent_cell.value:
                    adjacent_cell.font = font_green
                else:
                    adjacent_cell.font = font_red
            if cell.value == 'Исполнительные производства:':
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                if 'не было' in adjacent_cell.value:
                    adjacent_cell.font = font_green
                elif 'однако на момент отправки' in adjacent_cell.value:
                    adjacent_cell.font = font_orange
                else:
                    adjacent_cell.font = font_red
            if cell.value == 'Реестр недобросовестнных поставщиков:':
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                if 'не числится' in adjacent_cell.value:
                    adjacent_cell.font = font_green
                else:
                    adjacent_cell.font = font_red
            if cell.value in ['СВЕДЕНИЯ ИЗ РЕЕСТРА ЗАЛОГОВ', 'РОСФИНМОНИТОРИНГ(ТЕРРОРИСТИЧЕСКИЕ ОРГАНИЗАЦИИ)',
                              'САНКЦИИ США, ЕС, КАНАДЫ (ЮЛ)', 'СУЩЕСТВЕННЫЕ ФАКТЫ ПО ДАННЫМ ФЕДРЕСУРСА',
                              'ЧЕРНЫЙ СПИСОК ЦБ РФ (ЮЛ)', 'ПРОВЕРКА ПАСПОРТА', 'АРБИТРАЖНЫЙ СУД ФЛ',
                              'НАЛОГОВАЯ ЗАДОЛЖЕННОСТЬ (ФЛ)', 'ЗАБЛОКИРОВАННЫЕ РАСЧЕТНЫЕ СЧЕТА (ФЛ)',
                              'ФЕДЕРАЛЬНАЯ СЛУЖБА СУДЕБНЫХ ПРИСТАВОВ (ФССП ФЛ)', 'РЕЕСТР ЗАЛОГОВ',
                              'СООБЩЕНИЯ ФЕДРЕСУРС ФЛ', 'НАРКОКОНТРОЛЬ (АРХИВ ФСКН)', 'РОСФИНМОНИТОРИНГ: ЭКСТРЕМИСТЫ',
                              'САНКЦИИ США, ЕС, КАНАДЫ', 'ГАС РФ ПРАВОСУДИЕ', 'ОСОБЫЕ РЕЕСТРЫ',
                              'РЕЕСТР НЕДОБРОСОВЕСТНЫХ ПОСТАВЩИКОВ (РНП)', 'ФНС - СЕРВИС ПОЛУЧЕНИЯ ИНН',
                              'ДОЛЖНИКИ (ФЛ)', 'АРБИТРАЖНЫЙ СУД ИП/ЮЛ']:
                next_cell = sheet.cell(row=cell.row + 1, column=cell.column)
                if next_cell.value in ['Информация по данному источнику отсутствует', 'Информация не найдена',
                                       'Информация по источнику отсутствует',
                                       'Паспорт среди недействительных не значится'
                                       ] or 'По данным паспорта найден ИНН' in next_cell.value:
                    next_cell.font = font_green
                else:
                    next_cell.font = font_red
            if cell.value in ['Отсутствуют в реестре данные о контрагенте:',
                              'Находится ли контрагент в процессе ликвидации:',
                              'Находится ли контрагент в процессе банкротства:',
                              'Находится ли контрагент в процессе реорганизации:',
                              'Имеются ли в реестре отметки о недостоверности сведений:',
                              'Различается ли ИНН, ОГРН, адрес, директор, которые указаны в реестре, с теми, что предоставил контрагент:',
                              'Контрагент ведет деятельность на рынке менее 3 лет:',
                              'Имеются ли в отношении контрагента дела, в которых он выступает в качестве ответчика:',
                              'У компании имеются сообщения о банкротстве:',
                              'Имеются ли у контрагента исполнительные производства, которые в совокупности превышают 100 тыс. руб.:',
                              'Является ли деятельность контрагента убыточной:',
                              'Зарегистрирован ли контрагент по адресу массовой регистрации:',
                              'Является ли директор (учредители) контрагента массовым директором (учредителями):',
                              'Имеет ли организация превышающую 1000 рублей задолженность по уплате налогов и (или) не предоставляет налоговую отчетность:',
                              'Имеются ли у контрагента заблокированные расчетные счета:',
                              'Входит ли контрагент в перечень организаций и физических лиц, в отношении которых имеются сведения об их причастности к экстремистской деятельности или терроризму:',
                              'Контрагент числится в черном списке ЦБ РФ:',
                              'У контрагента отсутствуют государственные контракты:',
                              'Состоит ли контрагент в реестре недобросовестных поставщиков:']:
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                if adjacent_cell.value == 'Нет':
                    adjacent_cell.font = font_green
                else:
                    adjacent_cell.font = font_red
            if cell.value in ['Поставщик является заводом-изготовителем предмета лизинга:',
                              'Поставщик является официальным дилером завода-изготовителя предмета лизинга:',
                              'Имеются (имелись) ли у поставщика договоры купли-продажи с ООО «ЛКМБ-РТ»:']:
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                if adjacent_cell.value == 'Да':
                    adjacent_cell.font = font_green
                else:
                    adjacent_cell.font = font_red


def first_list(wb, name):
    sheet = wb['Лизингополучатель']
    check_date_of_registration(sheet)
    check_size_of_authorized_capital(sheet)
    delta_rating(sheet)
    activity_analysis(sheet)
    wb.save(name)


def second_list(wb, name):
    sheet = wb['Дир Учр Пор']
    activity_analysis(sheet)
    wb.save(name)


def third_list(wb, name):
    sheet = wb['Продавец']
    check_date_of_registration(sheet)
    check_size_of_authorized_capital(sheet)
    delta_rating(sheet)
    activity_analysis(sheet)
    wb.save(name)


def fourth_list(wb, name):
    sheet = wb['Предмет лизинга']
    wb.save(name)
