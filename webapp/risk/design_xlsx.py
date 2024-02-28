import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font


def main_design(file_name, individual=False):
    wb = openpyxl.load_workbook(file_name)

    font_title_name = Font(name="Century Gothic")
    alignment_wrap = Alignment(wrap_text=True)
    alignment_center_bold = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    alignment_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    alignment_top_left_bold = Alignment(
        horizontal="left", vertical="top", wrap_text=True
    )
    font_bold = Font(bold=True, name="Century Gothic")
    font_a1 = Font(bold=True, name="Century Gothic", size=19)
    font_a2 = Font(bold=True, name="Century Gothic", size=16)
    font_delta = Font(name="Century Gothic", size=28)
    color_a2 = PatternFill(start_color="ddeeff", end_color="ddeeff", fill_type="solid")
    color_delta = PatternFill(
        start_color="e9eef3", end_color="e9eef3", fill_type="solid"
    )

    if not individual:
        first_list(
            wb,
            file_name,
            font_title_name,
            alignment_wrap,
            font_a1,
            font_a2,
            alignment_center_bold,
            font_bold,
            color_a2,
            color_delta,
            alignment_top_left_bold,
            alignment_right,
            font_delta,
        )
        second_list(
            wb,
            file_name,
            font_title_name,
            alignment_wrap,
            font_a2,
            alignment_center_bold,
            font_bold,
            color_a2,
            color_delta,
            alignment_top_left_bold,
            alignment_right,
            font_delta,
        )
        third_list(
            wb,
            file_name,
            font_title_name,
            alignment_wrap,
            font_a2,
            alignment_center_bold,
            font_bold,
            color_a2,
            color_delta,
            alignment_top_left_bold,
            alignment_right,
            font_delta,
        )
        fourth_list(
            wb,
            file_name,
            font_title_name,
            alignment_wrap,
            font_a2,
            alignment_center_bold,
            font_bold,
            color_a2,
            color_delta,
            alignment_top_left_bold,
            alignment_right,
            font_delta,
        )
    else:
        second_list(
            wb,
            file_name,
            font_title_name,
            alignment_wrap,
            font_a2,
            alignment_center_bold,
            font_bold,
            color_a2,
            color_delta,
            alignment_top_left_bold,
            alignment_right,
            font_delta,
            individual=True,
        )


def overall_design(
    sheet,
    font_title_name,
    alignment_wrap,
    alignment_center_bold,
    font_bold,
    color_delta,
    alignment_top_left_bold,
    alignment_right,
    font_delta,
):
    sheet.column_dimensions["A"].width = 39
    sheet.column_dimensions["B"].width = 70
    sheet.column_dimensions["C"].width = 19
    sheet.column_dimensions["D"].width = 19
    for row in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row].height = None

    for cell in sheet["B"]:
        cell.alignment = alignment_wrap

    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font_title_name
            if cell.value in ["Наименование поля", "Значение", "Рейтинг", "Описание"]:
                cell.alignment = alignment_center_bold
                cell.font = font_bold
            if cell.value in [
                "Статус компании:",
                "Дата регистрации:",
                "Документы, поданные для регистрации/изменения в ЕГРЮЛ:",
                "Адрес регистрации организации:",
                "Размер уставного капитала:",
                "Массовость директоров/учредителей:",
                "Дисквалифицированные лица:",
                "Преступления, связанные с оборотом наркотиков по данным УФСКН России:",
                "Смена директоров/учредителей:",
                "Госконтракты:",
                "Сообщения о банкротстве:",
                "Налоговая задолженность и отчетность:",
                "Заблокированные расчетные счета:",
                "Арбитражные дела:",
                "Исполнительные производства:",
                "Наличие дочерних компаний:",
                "Реестр недобросовестнных поставщиков:",
            ]:
                cell.font = font_bold
                cell.alignment = alignment_wrap
            if cell.value in [
                "Анализ регистрационных данных",
                "Анализ директоров/учредителей",
                "Анализ деятельности",
            ]:
                cell.font = font_bold
                cell.fill = color_delta
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                adjacent_cell.fill = color_delta
            if cell.value in [
                "ДОЧЕРНИЕ ОРГАНИЗАЦИИ",
                "НАЛОГИ И СБОРЫ",
                "АФФИЛИРОВАННЫЕ И СВЯЗАННЫЕ ЛИЦА",
                "ПРОВЕРКА АФФИЛИРОВАННЫХ КОМПАНИЙ",
                "СУЩЕСТВЕННЫЕ ФАКТЫ ПО ДАННЫМ ФЕДРЕСУРСА",
                "СВЕДЕНИЯ ИЗ РЕЕСТРА ЗАЛОГОВ",
                "РОСФИНМОНИТОРИНГ(ТЕРРОРИСТИЧЕСКИЕ ОРГАНИЗАЦИИ)",
                "САНКЦИИ США, ЕС, КАНАДЫ (ЮЛ)",
                "ЧЕРНЫЙ СПИСОК ЦБ РФ (ЮЛ)",
                "СВЕДЕНИЯ ИЗ СМИ",
                "САЙТ В СЕТИ ИНТЕРНЕТ",
                "СОЦИАЛЬНЫЕ СЕТИ",
                "ЛИЗИНГ",
                "АНАЛИЗ ДОГОВОРОВ С КОНТРАГЕНТАМИ",
                "АРБИТРАЖНЫЕ ДЕЛА",
                "ПРОВЕРКА ПАСПОРТА",
                "Регистрация в качестве индивидуального предпринимателя:",
                "АРБИТРАЖНЫЙ СУД ФЛ",
                "АРБИТРАЖНЫЙ СУД ИП/ЮЛ",
                "НАЛОГОВАЯ ЗАДОЛЖЕННОСТЬ",
                "ЗАБЛОКИРОВАННЫЕ РАСЧЕТНЫЕ СЧЕТА (ФЛ)",
                "ФЕДЕРАЛЬНАЯ СЛУЖБА СУДЕБНЫХ ПРИСТАВОВ (ФССП ФЛ)",
                "РЕЕСТР ЗАЛОГОВ",
                "СООБЩЕНИЯ ФЕДРЕСУРС ФЛ",
                "НАРКОКОНТРОЛЬ (АРХИВ ФСКН)",
                "РОСФИНМОНИТОРИНГ: ЭКСТРЕМИСТЫ",
                "САНКЦИИ США, ЕС, КАНАДЫ",
                "ГАС РФ ПРАВОСУДИЕ",
                "ОСОБЫЕ РЕЕСТРЫ",
                "РЕЕСТР НЕДОБРОСОВЕСТНЫХ ПОСТАВЩИКОВ (РНП)",
                "ФНС - СЕРВИС ПОЛУЧЕНИЯ ИНН",
                "ДОЛЖНИКИ (ФЛ)",
                "Учредители ЮЛ",
                "Учредители ФЛ",
                "ПРОВЕРКА НА ДОЛЖНУЮ ОСТМОТРИТЕЛЬНОСТЬ",
                "ИСПОЛНИТЕЛЬНЫЕ ПРОИЗВОДСТВА",
                "АРБИТРАЖНЫЕ ДЕЛА",
                "НАИМЕНОВАНИЕ ПЛ",
                "ГОД ВЫПУСКА, СОСТОЯНИЕ",
                "СТОИМОСТЬ ПРЕДМЕТА ЛИЗИНГА",
                "ЛИКВИДНОСТЬ",
                "ПРАВОУСТАНАВЛИВАЮЩИЕ ДОКУМЕНТЫ",
                "ПРОВЕРКА ПО ДАННЫМ САЙТА ГИБДД/МИНСЕЛЬХОЗ",
                "ПРОВЕРКА ПО ДАННЫМ РЕЕСТРА ЗАЛОГОВ ФНП/ФЕДРЕСУРС",
                "ПРОВЕРКА ДЕЙСТВИТЕЛЬНОСТИ ЭПТС",
                "2.1. ДИРЕКТОР/ГЕН. ДИРЕКТОР",
                "2.2. УЧРЕДИТЕЛИ:",
                "2.3 ПОРУЧИТЕЛИ:",
            ]:
                cell.font = font_bold
            if cell.value in [
                "Информация о регистрирующем органе:",
                "Отсутствуют в реестре данные о контрагенте:",
                "Находится ли контрагент в процессе ликвидации:",
                "Находится ли контрагент в процессе банкротства:",
                "Находится ли контрагент в процессе реорганизации:",
                "Имеются ли в реестре отметки о недостоверности сведений:",
                "Различается ли ИНН, ОГРН, адрес, директор, которые указаны в реестре, с теми, что предоставил контрагент:",
                "Контрагент ведет деятельность на рынке менее 3 лет:",
                "Имеются ли в отношении контрагента дела, в которых он выступает в качестве ответчика:",
                "У компании имеются сообщения о банкротстве:",
                "Имеются ли у контрагента исполнительные производства, которые в совокупности превышают 100 тыс. руб.:",
                "Является ли деятельность контрагента убыточной:",
                "Зарегистрирован ли контрагент по адресу массовой регистрации:",
                "Является ли директор (учредители) контрагента массовым директором (учредителями):",
                "Имеет ли организация превышающую 1000 рублей задолженность по уплате налогов и (или) не предоставляет налоговую отчетность:",
                "Имеются ли у контрагента заблокированные расчетные счета:",
                "Входит ли контрагент в перечень организаций и физических лиц, в отношении которых имеются сведения об их причастности к экстремистской деятельности или терроризму:",
                "Контрагент числится в черном списке ЦБ РФ:",
                "У контрагента отсутствуют государственные контракты:",
                "Состоит ли контрагент в реестре недобросовестных поставщиков:",
                "Поставщик является заводом-изготовителем предмета лизинга:",
                "Поставщик является официальным дилером завода-изготовителя предмета лизинга:",
                "Имеются (имелись) ли у поставщика договоры купли-продажи с ООО «ЛКМБ-РТ»:",
            ]:
                cell.alignment = alignment_wrap
                adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                adjacent_cell.alignment = alignment_center_bold
            if cell.value in [
                "Является руководителем:",
                "Является учредителем:",
                "Являлся руководителем:",
                "Являлся учредителем:",
            ]:
                cell.alignment = alignment_center_bold
                cell.font = font_bold
                cell.fill = color_delta
                next_cell = sheet.cell(row=cell.row + 1, column=cell.column)
                next_cell.alignment = alignment_top_left_bold
            if isinstance(cell.value, str):
                if (
                    ("УЧРЕДИТЕЛЬ 1" in cell.value)
                    or ("УЧРЕДИТЕЛЬ 2" in cell.value)
                    or ("УЧРЕДИТЕЛЬ 3" in cell.value)
                    or ("УЧРЕДИТЕЛЬ 4" in cell.value)
                    or ("УЧРЕДИТЕЛЬ 5" in cell.value)
                ):
                    cell.font = font_bold
                    cell.fill = color_delta
                    adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                    adjacent_cell.fill = color_delta
            if isinstance(cell.value, str):
                if "ЧЕК-ЛИСТ" in cell.value:
                    cell.font = font_bold
            if cell.value in [
                "1. Сведения из ЕГРЮЛ",
                "2. Период деятельности",
                "3. Арбитражные дела",
                "4. Сведения о банкротстве",
                "5. Исполнительные производства",
                "6. Финансовая отчетность",
                "7. Адреса массовой регистрации",
                "8. «Массовость» руководителей и участников",
                "9. Налоговая задолженность",
                "10. Заблокированные счета",
                "11. Сведения о причастности к экстремистской деятельности по данным Росфинмониторинг",
                "12. Черный список ЦБ РФ",
                "13. Государственные контракты",
                "14. Реестр недобросовестных поставщиков",
                "15. Дилерство",
                "16. Кредитная история в ООО «ЛКМБ-РТ»",
            ]:
                cell.font = font_bold
                cell.fill = color_delta
                cell.alignment = alignment_center_bold
                sheet.merge_cells(
                    start_row=cell.row,
                    start_column=cell.column,
                    end_row=cell.row,
                    end_column=cell.column + 1,
                )


def first_list(
    wb,
    name,
    font_title_name,
    alignment_wrap,
    font_a1,
    font_a2,
    alignment_center_bold,
    font_bold,
    color_a2,
    color_delta,
    alignment_top_left_bold,
    alignment_right,
    font_delta,
):
    sheet = wb["Лизингополучатель"]
    overall_design(
        sheet,
        font_title_name,
        alignment_wrap,
        alignment_center_bold,
        font_bold,
        color_delta,
        alignment_top_left_bold,
        alignment_right,
        font_delta,
    )
    sheet["A1"].font = font_a1  # Риск-заключение по ООО "Ромашка"
    sheet["A2"].font = font_a2  # Названия глав (частей) риск-заключения
    sheet["A2"].fill = color_a2
    sheet["B2"].fill = color_a2
    for cell in sheet["A"]:
        if cell.value == "2. Анализ деятельности":
            cell.font = font_a2
            cell.fill = color_a2
            adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
            adjacent_cell.fill = color_a2
            next_cell = sheet.cell(row=cell.row + 1, column=cell.column)
            next_cell.alignment = alignment_center_bold
            next_cell.font = font_delta
            for num in range(3, 10):
                sheet.cell(row=cell.row + num, column=cell.column).alignment = (
                    alignment_center_bold
                )
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "ФИНАНСОВЫЕ ПОКАЗАТЕЛИ":
                next_cell = sheet.cell(row=cell.row + 1, column=cell.column)
                cell.font = font_bold
                cell.fill = color_delta
                cell.alignment = alignment_center_bold
                if (
                    next_cell.value == "Отсутствуют финансовые показатели"
                    or next_cell.value == "-"
                ):
                    sheet.merge_cells(
                        start_row=cell.row,
                        start_column=cell.column,
                        end_row=cell.row,
                        end_column=cell.column + 1,
                    )
                else:
                    sheet.merge_cells(
                        start_row=cell.row,
                        start_column=cell.column,
                        end_row=cell.row,
                        end_column=cell.column + 3,
                    )
                    next_cell.alignment = alignment_right
                    next_cell.font = font_bold
                    for num in range(1, 4):
                        sheet.cell(
                            row=cell.row + 1, column=cell.column + num
                        ).alignment = alignment_right
                        sheet.cell(row=cell.row + 1, column=cell.column + num).font = (
                            font_bold
                        )

                    for num in range(2, 8):
                        sheet.cell(row=cell.row + num, column=cell.column).alignment = (
                            alignment_right
                        )
                        sheet.cell(
                            row=cell.row + num, column=cell.column + 1
                        ).alignment = alignment_right
                        sheet.cell(
                            row=cell.row + num, column=cell.column + 2
                        ).alignment = alignment_right
                        sheet.cell(
                            row=cell.row + num, column=cell.column + 3
                        ).alignment = alignment_right
    wb.save(name)


def second_list(
    wb,
    name,
    font_title_name,
    alignment_wrap,
    font_a2,
    alignment_center_bold,
    font_bold,
    color_a2,
    color_delta,
    alignment_top_left_bold,
    alignment_right,
    font_delta,
    individual=False,
):
    if not individual:
        sheet = wb["Дир Учр Пор"]
    else:
        sheet = wb["Физ. лицо"]
    overall_design(
        sheet,
        font_title_name,
        alignment_wrap,
        alignment_center_bold,
        font_bold,
        color_delta,
        alignment_top_left_bold,
        alignment_right,
        font_delta,
    )
    sheet["A1"].font = (
        font_a2  # A1 к а2 потому что название Риск-заключение по ... встречается только на 1 листе
    )
    sheet["A1"].fill = color_a2
    sheet["B1"].fill = color_a2
    wb.save(name)


def third_list(
    wb,
    name,
    font_title_name,
    alignment_wrap,
    font_a2,
    alignment_center_bold,
    font_bold,
    color_a2,
    color_delta,
    alignment_top_left_bold,
    alignment_right,
    font_delta,
):
    sheet = wb["Продавец"]
    overall_design(
        sheet,
        font_title_name,
        alignment_wrap,
        alignment_center_bold,
        font_bold,
        color_delta,
        alignment_top_left_bold,
        alignment_right,
        font_delta,
    )
    sheet["A1"].font = (
        font_a2  # A1 к а2 потому что название Риск-заключение по ... встречается только на 1 листе
    )
    sheet["A1"].fill = color_a2
    sheet["B1"].fill = color_a2
    for cell in sheet["A"]:
        if cell.value == "2. Анализ деятельности":
            cell.font = font_a2
            cell.fill = color_a2
            adjacent_cell = sheet.cell(row=cell.row, column=cell.column + 1)
            adjacent_cell.fill = color_a2
            next_cell = sheet.cell(row=cell.row + 1, column=cell.column)
            next_cell.alignment = alignment_center_bold
            next_cell.font = font_delta
            for num in range(3, 10):
                sheet.cell(row=cell.row + num, column=cell.column).alignment = (
                    alignment_center_bold
                )
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "ФИНАНСОВЫЕ ПОКАЗАТЕЛИ":
                next_cell = sheet.cell(row=cell.row + 1, column=cell.column)
                cell.font = font_bold
                cell.fill = color_delta
                cell.alignment = alignment_center_bold
                if (
                    next_cell.value == "Отсутствуют финансовые показатели"
                    or next_cell.value == "-"
                ):
                    sheet.merge_cells(
                        start_row=cell.row,
                        start_column=cell.column,
                        end_row=cell.row,
                        end_column=cell.column + 1,
                    )
                else:
                    sheet.merge_cells(
                        start_row=cell.row,
                        start_column=cell.column,
                        end_row=cell.row,
                        end_column=cell.column + 3,
                    )
                    next_cell.alignment = alignment_right
                    next_cell.font = font_bold
                    for num in range(1, 4):
                        sheet.cell(
                            row=cell.row + 1, column=cell.column + num
                        ).alignment = alignment_right
                        sheet.cell(row=cell.row + 1, column=cell.column + num).font = (
                            font_bold
                        )

                    for num in range(2, 8):
                        sheet.cell(row=cell.row + num, column=cell.column).alignment = (
                            alignment_right
                        )
                        sheet.cell(
                            row=cell.row + num, column=cell.column + 1
                        ).alignment = alignment_right
                        sheet.cell(
                            row=cell.row + num, column=cell.column + 2
                        ).alignment = alignment_right
                        sheet.cell(
                            row=cell.row + num, column=cell.column + 3
                        ).alignment = alignment_right

    wb.save(name)


def fourth_list(
    wb,
    name,
    font_title_name,
    alignment_wrap,
    font_a2,
    alignment_center_bold,
    font_bold,
    color_a2,
    color_delta,
    alignment_top_left_bold,
    alignment_right,
    font_delta,
):
    sheet = wb["Предмет лизинга"]
    overall_design(
        sheet,
        font_title_name,
        alignment_wrap,
        alignment_center_bold,
        font_bold,
        color_delta,
        alignment_top_left_bold,
        alignment_right,
        font_delta,
    )
    sheet["A1"].font = (
        font_a2  # A1 к а2 потому что название Риск-заключение по ... встречается только на 1 листе
    )
    sheet["A1"].fill = color_a2
    sheet["B1"].fill = color_a2
    wb.save(name)
