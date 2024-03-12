import openpyxl

from datetime import datetime as dt
from flask_login import current_user

from webapp.risk.design_xlsx import main_design
from webapp.risk.conditions_xlsx import main_conditions
from webapp.risk.logger import logging
from webapp.risk.models import Okved
from webapp.config import PATH_FOR_HTML_PAGES, PATH_FOR_HTML_PAGES_IND


def create_xlsx_file_individual(delta_info: dict, person):
    def write_user(data):
        for index, (key, value) in enumerate(data.items()):
            if index == 4:  # Кол-во итераций
                break
            sheet[f"A{sheet.max_row + 1}"].value = key
            sheet[f"B{sheet.max_row}"].value = value

        for _ in range(20):
            sheet[f"A{sheet.max_row + 1}"].value = ""
        sheet[f"A{sheet.max_row + 1}"].value = "ПРОВЕРКА ПАСПОРТА"
        sheet[f"A{sheet.max_row + 1}"].value = (
            "Паспорт среди недействительных не значится"
        )
        for _ in range(20):
            sheet[f"A{sheet.max_row + 1}"].value = ""

        sheet[f"A{sheet.max_row + 2}"].value = (
            "Регистрация в качестве индивидуального предпринимателя:"
        )
        if data["Инфо_ИП"]:
            for num in data["Инфо_ИП"]:
                for key, value in data["Инфо_ИП"][num].items():
                    sheet[f"A{sheet.max_row + 1}"].value = key
                    sheet[f"B{sheet.max_row}"].value = value
                sheet[f"A{sheet.max_row + 1}"].value = ""
        else:
            sheet[f"A{sheet.max_row + 1}"].value = "Информация отсутствует"
            sheet[f"A{sheet.max_row + 1}"].value = ""

        sheet[f"A{sheet.max_row + 1}"].value = "Адрес регистрации:"
        sheet[f"B{sheet.max_row}"].value = data["Адрес_регистрации"]

        sheet[f"A{sheet.max_row + 2}"].value = "Информация о регистрирующем органе:"
        sheet[f"B{sheet.max_row}"].value = data["Имя_налог_органа"]

        sheet[f"A{sheet.max_row + 2}"].value = "Является руководителем:"
        sheet[f"B{sheet.max_row}"].value = "Является учредителем:"
        sheet[f"A{sheet.max_row + 1}"].value = data["История_руководства"][
            "Является руководителем"
        ]
        sheet[f"B{sheet.max_row}"].value = data["История_руководства"][
            "Является учредителем"
        ]
        sheet[f"A{sheet.max_row + 1}"].value = "Являлся руководителем:"
        sheet[f"B{sheet.max_row}"].value = "Являлся учредителем:"
        sheet[f"A{sheet.max_row + 1}"].value = data["История_руководства"][
            "Являлся руководителем"
        ]
        sheet[f"B{sheet.max_row}"].value = data["История_руководства"][
            "Являлся учредителем"
        ]

        for index, (key, value) in enumerate(data.items()):
            if index == 24:  # Кол-во итераций
                break
            if 13 > index >= 9:
                sheet[f"A{sheet.max_row + 2}"].value = key
                sheet[f"A{sheet.max_row + 1}"].value = value
            if 16 > index >= 13:
                sheet[f"A{sheet.max_row + 2}"].value = key
                sheet[f"A{sheet.max_row + 1}"].value = value
                for _ in range(20):
                    sheet[f"A{sheet.max_row + 1}"].value = ""
            if index >= 16:
                sheet[f"A{sheet.max_row + 2}"].value = key
                sheet[f"A{sheet.max_row + 1}"].value = value

        sheet[f"A{sheet.max_row + 2}"].value = "СОЦИАЛЬНЫЕ СЕТИ"
        for _ in range(3):
            sheet[f"A{sheet.max_row + 1}"].value = ""

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet["A1"].value = (
        f'Анализ физ. лица {person.get_full_name} на {dt.today().strftime("%d.%m.%Y")}'
    )
    try:
        sheet_name1 = wb.sheetnames[0]  # берем название первого листа
        wb[sheet_name1].title = "Физ. лицо"  # дали новое название листу
    except Exception as _ex:
        logging.info(_ex, exc_info=True)

    sheet[f"A{sheet.max_row + 1}"].value = "1. Общее описание Физ. лица"
    sheet[f"A{sheet.max_row + 1}"].value = person.get_full_name
    sheet[f"A{sheet.max_row + 1}"].value = ""

    write_user(delta_info)

    logging.info(f"Сохраняем файл. Created by {current_user}")
    individual_path = rf"{PATH_FOR_HTML_PAGES_IND}/{person.get_full_name} ИНН {person.get_inn}/{dt.today().strftime('%d.%m.%Y')}"
    wb.save(f"{individual_path}/Анализ физ. лица {person.get_full_name}.xlsx")

    logging.info(f"({current_user}). Запускаем оформеление дизайна xlsx")
    main_design(f"{individual_path}/Анализ физ. лица {person.get_full_name}.xlsx", True)
    main_conditions(
        f"{individual_path}/Анализ физ. лица {person.get_full_name}.xlsx", True
    )


def create_xlsx_file(
    inn_client,
    inn_seller,
    main_client: dict,
    delta_client: dict,
    director_client: dict,
    founders_client: dict,
    main_seller: dict,
    delta_seller: dict,
    last_table_seller: dict,
    short_name: str,
):
    """main_client: can't be None
    delta_client: can be None, если клиент ИП/КФХ/физик
    director_client: can be None, если клиент ИП/КФХ/физик
    founders_client: can be None, если клиент ИП/КФХ/физик
    main_seller: can be None, если Возвратный
    delta_seller: can be None, если Возвратный/если клиент ИП/КФХ/физик
    last_table_seller: can be None, если Возвратный"""

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet["A1"].value = (
        f'Риск заключение по {main_client["Краткое наименование"]} на {dt.today().strftime("%d.%m.%Y")}'
    )
    try:
        sheet_name1 = wb.sheetnames[0]  # берем название первого листа
        wb.create_sheet(title="Дир Учр Пор")
        wb.create_sheet(title="Продавец")
        wb.create_sheet(title="Предмет лизинга")
        wb[sheet_name1].title = "Лизингополучатель"  # дали новое название первому листу
    except Exception as _ex:
        logging.info(_ex, exc_info=True)

    def write_user(data):
        for index, (key, value) in enumerate(data.items()):
            if index == 4:  # Кол-во итераций
                break
            sheet[f"A{sheet.max_row + 1}"].value = key
            sheet[f"B{sheet.max_row}"].value = value

        for _ in range(20):
            sheet[f"A{sheet.max_row + 1}"].value = ""
        sheet[f"A{sheet.max_row + 1}"].value = "ПРОВЕРКА ПАСПОРТА"
        sheet[f"A{sheet.max_row + 1}"].value = (
            "Паспорт среди недействительных не значится"
        )
        for _ in range(20):
            sheet[f"A{sheet.max_row + 1}"].value = ""

        sheet[f"A{sheet.max_row + 2}"].value = (
            "Регистрация в качестве индивидуального предпринимателя:"
        )
        if data["Инфо_ИП"]:
            for num in data["Инфо_ИП"]:
                for key, value in data["Инфо_ИП"][num].items():
                    sheet[f"A{sheet.max_row + 1}"].value = key
                    sheet[f"B{sheet.max_row}"].value = value
                sheet[f"A{sheet.max_row + 1}"].value = ""
        else:
            sheet[f"A{sheet.max_row + 1}"].value = "Информация отсутствует"
            sheet[f"A{sheet.max_row + 1}"].value = ""

        sheet[f"A{sheet.max_row + 1}"].value = "Адрес регистрации:"
        sheet[f"B{sheet.max_row}"].value = data["Адрес_регистрации"]

        sheet[f"A{sheet.max_row + 2}"].value = "Информация о регистрирующем органе:"
        sheet[f"B{sheet.max_row}"].value = data["Имя_налог_органа"]

        sheet[f"A{sheet.max_row + 2}"].value = "Является руководителем:"
        sheet[f"B{sheet.max_row}"].value = "Является учредителем:"
        sheet[f"A{sheet.max_row + 1}"].value = data["История_руководства"][
            "Является руководителем"
        ]
        sheet[f"B{sheet.max_row}"].value = data["История_руководства"][
            "Является учредителем"
        ]
        sheet[f"A{sheet.max_row + 1}"].value = "Являлся руководителем:"
        sheet[f"B{sheet.max_row}"].value = "Являлся учредителем:"
        sheet[f"A{sheet.max_row + 1}"].value = data["История_руководства"][
            "Являлся руководителем"
        ]
        sheet[f"B{sheet.max_row}"].value = data["История_руководства"][
            "Являлся учредителем"
        ]

        for index, (key, value) in enumerate(data.items()):
            if index == 24:  # Кол-во итераций
                break
            if 13 > index >= 9:
                sheet[f"A{sheet.max_row + 2}"].value = key
                sheet[f"A{sheet.max_row + 1}"].value = value
            if 16 > index >= 13:
                sheet[f"A{sheet.max_row + 2}"].value = key
                sheet[f"A{sheet.max_row + 1}"].value = value
                for _ in range(20):
                    sheet[f"A{sheet.max_row + 1}"].value = ""
            if index >= 16:
                sheet[f"A{sheet.max_row + 2}"].value = key
                sheet[f"A{sheet.max_row + 1}"].value = value

        sheet[f"A{sheet.max_row + 2}"].value = "СОЦИАЛЬНЫЕ СЕТИ"
        for _ in range(3):
            sheet[f"A{sheet.max_row + 1}"].value = ""

    def write_company(data, data_delta):
        sheet[f"A{sheet.max_row + 1}"].value = "Наименование поля"
        sheet[f"B{sheet.max_row}"].value = "Значение"

        for index, (name, value) in enumerate(data.items()):
            if index == 22:  # Кол-во итераций. Заканчивается на "ИСТОРИЯ"
                break
            match value:
                case str() | Okved():
                    sheet[f"A{sheet.max_row + 1}"].value = name
                    sheet[f"B{sheet.max_row}"].value = value
                case dict():
                    some_list = []
                    try:
                        for num in data.get(name):
                            some_list.append(
                                f'{num}) {data.get(name).get(num).get("percent")} {data.get(name).get(num).get("sum")} {data.get(name).get(num).get("full_name")} {data.get(name).get(num).get("inn")} {data.get(name).get(num).get("egrul")}'
                            )
                    except TypeError:
                        logging.info("Попал в except")
                    sheet[f"A{sheet.max_row + 1}"].value = name
                    if len(some_list) > 0:
                        sheet[f"B{sheet.max_row}"].value = "\n".join(some_list)
                    else:
                        sheet[f"B{sheet.max_row}"].value = "-"
                case _:
                    logging.info(type(value))
                    logging.info(value)
                    logging.info(f"{index} {name}: {value} (other)")
                    sheet[f"A{sheet.max_row + 1}"].value = name
                    sheet[f"B{sheet.max_row}"].value = "-"

        sheet[f"A{sheet.max_row + 2}"].value = "2. Анализ деятельности"
        sheet[f"A{sheet.max_row + 1}"].value = data_delta["Рейтинг дельта номер"]
        sheet[f"B{sheet.max_row}"].value = data_delta["Рейтинг дельта текст"]
        sheet[f"A{sheet.max_row + 1}"].value = "Рейтинг"
        sheet[f"B{sheet.max_row}"].value = "Описание"
        delta_rating_dict = {
            "100": "Благонадежность  предприятия очень высокая.",
            "90": "Вероятность благонадежности предприятия высокая.",
            "80": "У предприятия имеется один или несколько признаков неблагонадежности.",
            "70": "У предприятия имеется несколько признаков неблагонадежности",
            "50-60": "Взаимодействие с данной фирмой может быть сопряжено со средней степенью риска.",
            "40": "Взаимодействие с данной фирмой может быть сопряжено с высокой степенью риска.",
            "0-30": "Сотрудничество с данной компанией нежелательно",
        }

        for key, value in delta_rating_dict.items():
            sheet[f"A{sheet.max_row + 1}"].value = key
            sheet[f"B{sheet.max_row}"].value = value

        sheet[f"A{sheet.max_row + 2}"].value = "Анализ регистрационных данных"
        for index, (key, value) in enumerate(data_delta.items()):
            if index == 5:
                break
            sheet[f"A{sheet.max_row + 1}"].value = key
            sheet[f"B{sheet.max_row}"].value = value

        sheet[f"A{sheet.max_row + 2}"].value = "Анализ директоров/учредителей"
        for index, (key, value) in enumerate(data_delta.items()):
            if index == 9:
                break
            if index >= 5:
                sheet[f"A{sheet.max_row + 1}"].value = key
                sheet[f"B{sheet.max_row}"].value = value

        sheet[f"A{sheet.max_row + 2}"].value = "Анализ деятельности"
        for index, (key, value) in enumerate(data_delta.items()):
            if index == 17:
                break
            if index >= 9:
                sheet[f"A{sheet.max_row + 1}"].value = key
                sheet[f"B{sheet.max_row}"].value = value

        sheet[f"A{sheet.max_row + 2}"].value = "ФИНАНСОВЫЕ ПОКАЗАТЕЛИ"
        columns_list = ["B", "C", "D"]
        columns_list_index = 0
        num_for_values = -5

        try:
            if data["Финансы"].get(2022):
                sheet[f"A{sheet.max_row + 1}"].value = "НАИМЕНОВАНИЕ"
                sheet[f"A{sheet.max_row + 1}"].value = "Баланс"
                sheet[f"A{sheet.max_row + 1}"].value = "Выручка"
                sheet[f"A{sheet.max_row + 1}"].value = "Обязательства"
                sheet[f"A{sheet.max_row + 1}"].value = "Чистая прибыль"
                sheet[f"A{sheet.max_row + 1}"].value = "Капитал и резервы"
                sheet[f"A{sheet.max_row + 1}"].value = "Основные средства"
                for key in data["Финансы"]:
                    sheet[
                        f"{columns_list[columns_list_index]}{sheet.max_row - 6}"
                    ].value = key
                    for indicators, values in data["Финансы"][key].items():
                        sheet[
                            f"{columns_list[columns_list_index]}{sheet.max_row + num_for_values}"
                        ].value = values
                        num_for_values += 1
                    num_for_values = -5
                    columns_list_index += 1
            else:
                sheet[f"A{sheet.max_row + 1}"].value = (
                    "Отсутствуют финансовые показатели"
                )
        except Exception as _ex:
            logging.info(_ex, exc_info=True)
            sheet[f"A{sheet.max_row + 1}"].value = "-"

        sheet[f"A{sheet.max_row + 2}"].value = "ДОЧЕРНИЕ ОРГАНИЗАЦИИ"
        sheet[f"B{sheet.max_row}"].value = data["Дочерние организации"]

        sheet[f"A{sheet.max_row + 2}"].value = "НАЛОГИ И СБОРЫ"
        sheet[f"B{sheet.max_row}"].value = data["Налоги и сборы"]

        sheet[f"A{sheet.max_row + 2}"].value = "АФФИЛИРОВАННЫЕ И СВЯЗАННЫЕ ЛИЦА"
        for _ in range(20):
            sheet[f"A{sheet.max_row + 1}"].value = ""

        sheet[f"A{sheet.max_row + 2}"].value = "ПРОВЕРКА АФФИЛИРОВАННЫХ КОМПАНИЙ"
        for _ in range(20):
            sheet[f"A{sheet.max_row + 1}"].value = ""

        sheet[f"A{sheet.max_row + 2}"].value = "СУЩЕСТВЕННЫЕ ФАКТЫ ПО ДАННЫМ ФЕДРЕСУРСА"
        sheet[f"A{sheet.max_row + 1}"].value = data["Федресурс"]
        for _ in range(20):
            sheet[f"A{sheet.max_row + 1}"].value = ""

        sheet[f"A{sheet.max_row + 2}"].value = "СВЕДЕНИЯ ИЗ РЕЕСТРА ЗАЛОГОВ"
        sheet[f"A{sheet.max_row + 1}"].value = data["Реестр залогов"]
        for _ in range(20):
            sheet[f"A{sheet.max_row + 1}"].value = ""

        sheet[f"A{sheet.max_row + 2}"].value = (
            "РОСФИНМОНИТОРИНГ(ТЕРРОРИСТИЧЕСКИЕ ОРГАНИЗАЦИИ)"
        )
        sheet[f"A{sheet.max_row + 1}"].value = data["Росфинмониторинг"]

        sheet[f"A{sheet.max_row + 2}"].value = "САНКЦИИ США, ЕС, КАНАДЫ (ЮЛ)"
        sheet[f"A{sheet.max_row + 1}"].value = data["Санкции"]

        sheet[f"A{sheet.max_row + 2}"].value = "ЧЕРНЫЙ СПИСОК ЦБ РФ (ЮЛ)"
        sheet[f"A{sheet.max_row + 1}"].value = data["Черный список"]

        sheet[f"A{sheet.max_row + 2}"].value = "СВЕДЕНИЯ ИЗ СМИ"
        sheet[f"A{sheet.max_row + 1}"].value = (
            "Информация по данному источнику отсутствует"
        )

        sheet[f"A{sheet.max_row + 2}"].value = "САЙТ В СЕТИ ИНТЕРНЕТ"
        sheet[f"A{sheet.max_row + 1}"].value = (
            "Информация по данному источнику отсутствует"
        )

        sheet[f"A{sheet.max_row + 2}"].value = "СОЦИАЛЬНЫЕ СЕТИ"
        sheet[f"A{sheet.max_row + 1}"].value = (
            "Информация по данному источнику отсутствует"
        )

    sheet[f"A{sheet.max_row + 1}"].value = "1. Общее описание Лизингополучателя"

    """Если клиент Юр. лицо, то попадаем в if"""
    if len(inn_client) == 10:
        logging.info("Клиент юр. лицо. Идет заполнение файла...")
        write_company(main_client, delta_client)
        sheet[f"A{sheet.max_row + 2}"].value = "ЛИЗИНГ"
        for _ in range(20):
            sheet[f"A{sheet.max_row + 1}"].value = ""

        sheet[f"A{sheet.max_row + 2}"].value = "АНАЛИЗ ДОГОВОРОВ С КОНТРАГЕНТАМИ"
        for _ in range(20):
            sheet[f"A{sheet.max_row + 1}"].value = ""

        sheet[f"A{sheet.max_row + 2}"].value = "АРБИТРАЖНЫЕ ДЕЛА"
        for _ in range(20):
            sheet[f"A{sheet.max_row + 1}"].value = ""

        sheet[f"A{sheet.max_row + 2}"].value = "ИСПОЛНИТЕЛЬНЫЕ ПРОИЗВОДСТВА"

        logging.info(
            "Запуск процесса записи информации про директора/учредителей в xslx файл"
        )
        sheet = wb["Дир Учр Пор"]
        sheet[f"A{sheet.max_row}"].value = (
            "2. Анализ директора/учредителя (ей) / поручителя(ей)"
        )
        sheet[f"A{sheet.max_row + 1}"].value = "2.1. ДИРЕКТОР/ГЕН. ДИРЕКТОР"
        sheet[f"A{sheet.max_row + 1}"].value = director_client["Краткое наименование"]
        sheet[f"A{sheet.max_row + 1}"].value = ""

        write_user(director_client)

        sheet[f"A{sheet.max_row + 2}"].value = "2.2. УЧРЕДИТЕЛИ:"
        sheet[f"A{sheet.max_row + 1}"].value = "Учредители ЮЛ"

        founders_list_ul = []
        name_ul = "УЧРЕДИТЕЛИ ЮЛ"
        try:
            for num in main_client.get(name_ul):
                founders_list_ul.append(
                    f'{num}) {main_client.get(name_ul).get(num).get("percent")} {main_client.get(name_ul).get(num).get("sum")} {main_client.get(name_ul).get(num).get("full_name")} {main_client.get(name_ul).get(num).get("inn")} {main_client.get(name_ul).get(num).get("egrul")}'
                )
        except TypeError as _ex:
            logging.info("Попал в except")
            logging.info(_ex, exc_info=True)
        if len(founders_list_ul) > 0:
            sheet[f"B{sheet.max_row}"].value = "\n".join(founders_list_ul)
        else:
            sheet[f"B{sheet.max_row}"].value = "-"

        sheet[f"A{sheet.max_row + 2}"].value = "Учредители ФЛ"

        founders_list_fl = []
        name_fl = "УЧРЕДИТЕЛИ ФЛ"
        try:
            for num in main_client.get(name_fl):
                founders_list_fl.append(
                    f'{num}) {main_client.get(name_fl).get(num).get("percent")} {main_client.get(name_fl).get(num).get("sum")} {main_client.get(name_fl).get(num).get("full_name")} {main_client.get(name_fl).get(num).get("inn")} {main_client.get(name_fl).get(num).get("egrul")}'
                )
        except TypeError as _ex:
            logging.info("Попал в except")
            logging.info(_ex, exc_info=True)
        if len(founders_list_fl) > 0:
            sheet[f"B{sheet.max_row}"].value = "\n".join(founders_list_fl)
        else:
            sheet[f"B{sheet.max_row}"].value = "-"

        try:
            if main_client["УЧРЕДИТЕЛИ ФЛ"]:
                for num in founders_client:
                    if len(founders_client[num]) > 2:
                        sheet[f"A{sheet.max_row + 1}"].value = ""
                        sheet[f"A{sheet.max_row + 1}"].value = (
                            f'УЧРЕДИТЕЛЬ {num} {founders_client[num]["percent"]}'
                        )
                        sheet[f"A{sheet.max_row + 1}"].value = founders_client[num][
                            "full_name"
                        ]
                        sheet[f"A{sheet.max_row + 1}"].value = ""
                        write_user(founders_client[num])
                    else:
                        sheet[f"A{sheet.max_row + 1}"].value = ""
                        sheet[f"A{sheet.max_row + 1}"].value = (
                            f'УЧРЕДИТЕЛЬ {num} {founders_client[num]["percent"]} (ДИРЕКТОР)'
                        )
                        sheet[f"A{sheet.max_row + 1}"].value = founders_client[num][
                            "full_name"
                        ]
                        sheet[f"A{sheet.max_row + 1}"].value = (
                            "Проверка директора уже проведена"
                        )

        except Exception as _ex:
            logging.info(_ex, exc_info=True)

        sheet[f"A{sheet.max_row + 2}"].value = "2.3 ПОРУЧИТЕЛИ:"
        sheet[f"A{sheet.max_row + 1}"].value = (
            "Поручителями является директор/учредитель (проверка уже проведена)"
        )

    else:
        logging.info("Клиент ИП/КФХ. Идет заполнение файла...")
        sheet[f"A{sheet.max_row + 1}"].value = main_client["Краткое наименование"]
        sheet[f"A{sheet.max_row + 1}"].value = ""
        write_user(main_client)
        sheet = wb["Дир Учр Пор"]
        sheet[f"A{sheet.max_row}"].value = (
            "2. Анализ директора/учредителя (ей)/поручителя(ей)"
        )
        sheet[f"A{sheet.max_row + 1}"].value = (
            "Лизингополучатель ИП/КФХ, проверка уже проведена"
        )

    logging.info("Заполнение информации о продавце")
    sheet = wb["Продавец"]
    if inn_client != inn_seller:
        sheet[f"A{sheet.max_row}"].value = (
            f'3. Анализ продавца {main_seller["Краткое наименование"]}'
        )
    else:
        sheet[f"A{sheet.max_row}"].value = (
            f'3. Анализ продавца {main_client["Краткое наименование"]}'
        )
    try:
        if inn_client != inn_seller:
            if len(inn_seller) == 10:
                logging.info("Продавец юр. лицо")
                write_company(main_seller, delta_seller)
                sheet[f"A{sheet.max_row + 1}"].value = ""
            else:
                logging.info("Продавец ИП/КФХ")
                sheet[f"A{sheet.max_row + 1}"].value = main_seller[
                    "Краткое наименование"
                ]
                sheet[f"A{sheet.max_row + 1}"].value = ""
                write_user(main_seller)

            sheet[f"A{sheet.max_row + 2}"].value = (
                f'ЧЕК-ЛИСТ {main_seller["Краткое наименование"]}'
            )
            for key, value in last_table_seller.items():
                sheet[f"A{sheet.max_row + 1}"].value = key
                sheet[f"B{sheet.max_row}"].value = value

            sheet[f"A{sheet.max_row + 2}"].value = "ИСПОЛНИТЕЛЬНЫЕ ПРОИЗВОДСТВА"
            for _ in range(20):
                sheet[f"A{sheet.max_row + 1}"].value = ""
            sheet[f"A{sheet.max_row + 2}"].value = "АРБИТРАЖНЫЕ ДЕЛА"
            for _ in range(20):
                sheet[f"A{sheet.max_row + 1}"].value = ""
            sheet[f"A{sheet.max_row + 3}"].value = (
                "ПРОВЕРКА НА ДОЛЖНУЮ ОСМОТРИТЕЛЬНОСТЬ"
            )

        else:
            sheet[f"A{sheet.max_row + 1}"].value = (
                "Возвратный лизинг. Проверка лизингополучателя уже проведена"
            )
    except Exception as _ex:
        logging.info(_ex, exc_info=True)

    sheet = wb["Предмет лизинга"]
    sheet[f"A{sheet.max_row}"].value = "4. Анализ предмета лизинга"
    sheet[f"A{sheet.max_row + 2}"].value = "НАИМЕНОВАНИЕ ПЛ"
    sheet[f"A{sheet.max_row + 1}"].value = "ГОД ВЫПУСКА, СОСТОЯНИЕ"
    for _ in range(20):
        sheet[f"A{sheet.max_row + 1}"].value = ""
    sheet[f"A{sheet.max_row + 1}"].value = "СТОИМОСТЬ ПРЕДМЕТА ЛИЗИНГА"
    sheet[f"A{sheet.max_row + 1}"].value = ""
    sheet[f"A{sheet.max_row + 1}"].value = (
        "Стоимость предмета лизинга соответствует среднерыночной цене"
    )
    sheet[f"A{sheet.max_row + 1}"].value = (
        "(на основании сравнительного анализа аналогичного имущества в общедоступных источниках)"
    )
    sheet[f"A{sheet.max_row + 2}"].value = "ЛИКВИДНОСТЬ"
    sheet[f"A{sheet.max_row + 1}"].value = "Высокая/Средняя/Низкая/Безнадежная"
    for _ in range(7):
        sheet[f"A{sheet.max_row + 1}"].value = ""
    sheet[f"A{sheet.max_row + 1}"].value = "ПРАВОУСТАНАВЛИВАЮЩИЕ ДОКУМЕНТЫ"
    sheet[f"A{sheet.max_row + 1}"].value = "1. Договор купли-продажи"
    sheet[f"B{sheet.max_row}"].value = "Да/Нет (ПЛ новый)/Нет"
    sheet[f"A{sheet.max_row + 1}"].value = "2. Акт примема-передачи"
    sheet[f"B{sheet.max_row}"].value = "Да/Нет (ПЛ новый)/Нет"
    sheet[f"A{sheet.max_row + 1}"].value = "3. ПТС/ПСМ"
    sheet[f"B{sheet.max_row}"].value = "Да/Нет (ПЛ новый)/Нет"
    sheet[f"A{sheet.max_row + 2}"].value = "ПРОВЕРКА ПО ДАННЫМ САЙТА ГИБДД/МИНСЕЛЬХОЗ"
    sheet[f"A{sheet.max_row + 1}"].value = (
        "Существенная негативная информация не обнаружена/обнаружена"
    )
    for _ in range(20):
        sheet[f"A{sheet.max_row + 1}"].value = ""
    sheet[f"A{sheet.max_row + 2}"].value = (
        "ПРОВЕРКА ПО ДАННЫМ РЕЕСТРА ЗАЛОГОВ ФНП/ФЕДРЕСУРС"
    )
    sheet[f"A{sheet.max_row + 1}"].value = (
        "Существенная негативная информация не обнаружена/обнаружена"
    )
    for _ in range(20):
        sheet[f"A{sheet.max_row + 1}"].value = ""
    sheet[f"A{sheet.max_row + 1}"].value = "ПРОВЕРКА ДЕЙСТВИТЕЛЬНОСТИ ЭПТС"
    logging.info(f"Сохраняем файл. Created by {current_user}")
    wb.save(
        rf"{PATH_FOR_HTML_PAGES}/{short_name} ИНН {inn_client}/{dt.today().strftime('%d.%m.%Y')}/Риск заключение {inn_client}.xlsx"
    )

    logging.info(f"({current_user}). Запускаем оформеление дизайна xlsx")
    main_design(
        rf"{PATH_FOR_HTML_PAGES}/{short_name} ИНН {inn_client}/{dt.today().strftime('%d.%m.%Y')}/Риск заключение {inn_client}.xlsx"
    )
    main_conditions(
        rf"{PATH_FOR_HTML_PAGES}/{short_name} ИНН {inn_client}/{dt.today().strftime('%d.%m.%Y')}/Риск заключение {inn_client}.xlsx"
    )
