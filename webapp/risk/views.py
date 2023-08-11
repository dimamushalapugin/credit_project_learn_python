import time

import openpyxl
import os
from flask import Blueprint, flash, render_template, redirect, request, url_for, send_from_directory
from webapp.user.auth_utils import admin_required

blueprint = Blueprint('risk', __name__, url_prefix='/risk')


@blueprint.route('/risk_conclusion')
@admin_required
def risk_page():
    print('__risk_page__')
    filenames = os.listdir('webapp/static/files')
    return render_template('risk_conclusion.html', filenames=filenames)


@blueprint.route('/risk_conclusion/<path:folder_path>')
def risk_conclusion_folder(folder_path, filename):
    folder_path = os.path.join('webapp/static/files', folder_path)
    if os.path.exists(folder_path) and os.path.isdir(folder_path):
        items = os.listdir(folder_path)
        return render_template('risk_conclusion_folder.html', items=items)
    else:
        return "Folder not found", 404


def create_xlsx_file(data):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = data['client_inn']
    ws['B1'] = data['seller_inn']

    file_name = 'test.xlsx'
    full_file_path = os.path.join('webapp', 'static', 'files', file_name)
    wb.save(full_file_path)

    return file_name


@blueprint.route('/create_xlsx', methods=['POST'])
def create_xlsx():
    try:
        data = request.form
        time.sleep(2)
        file_name = create_xlsx_file(data)
        flash(f'Файл успешно создан', 'success')
        return redirect(url_for('risk.risk_page', file_name=file_name))
    except Exception as e:
        flash(str(e), 'warning')
        return redirect(url_for('risk.risk_page'))


@blueprint.route('/download/<path:filename>')
def download(filename):
    return send_from_directory(f'static/files/', filename, as_attachment=True)
